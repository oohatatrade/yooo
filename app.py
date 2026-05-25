from __future__ import annotations

import json
import mimetypes
import os
import posixpath
import threading
import time
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote, unquote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from akita_scraper import SearchParams, scrape_contracts


ROOT = Path(__file__).resolve().parent
PUBLIC_DIR = ROOT / "public"
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


def write_workbook(rows: list[dict], params: SearchParams) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = "akita_orders" if params.info_type == "order" else "akita_contracts"
    filename = f"{prefix}_{timestamp}.xlsx"
    path = OUTPUT_DIR / filename

    wb = Workbook()
    ws = wb.active
    if params.info_type == "order":
        ws.title = "発注情報"
        headers = [
            ("公開日", "published_date"),
            ("入札方式", "bidding_method"),
            ("工事・委託名称", "project_name"),
            ("工事・委託場所", "location"),
            ("工事・業務種別", "work_type"),
            ("等級", "grade"),
            ("工事・委託概要", "summary"),
            ("予定価格", "estimated_price"),
            ("入札執行課所", "department"),
            ("調達区分", "procurement_category"),
            ("詳細 URL", "detail_url"),
            ("検索キーワード", "matched_keyword"),
            ("取得日時", "scraped_at"),
        ]
        widths = [12, 14, 42, 34, 28, 10, 50, 18, 18, 12, 62, 22, 20]
    else:
        ws.title = "契約結果"
        headers = [
            ("契約公表番号", "contract_no"),
            ("入札方式", "bidding_method"),
            ("工事・委託名称", "project_name"),
            ("工事・委託場所", "location"),
            ("工事・委託概要", "summary"),
            ("工事・業務種別", "work_type"),
            ("請負者", "contractor"),
            ("県内・県外", "contractor_area"),
            ("契約金額", "contract_amount"),
            ("開札日", "bid_open_date"),
            ("契約日", "contract_date"),
            ("公表課所", "department"),
            ("調達区分", "procurement_category"),
            ("PDF URL", "pdf_url"),
            ("検索キーワード", "matched_keyword"),
            ("取得日時", "scraped_at"),
        ]
        widths = [18, 14, 36, 28, 46, 28, 28, 12, 14, 12, 12, 18, 12, 62, 22, 20]

    ws.append([label for label, _ in headers])
    for row in rows:
        ws.append([row.get(key, "") for _, key in headers])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    meta = wb.create_sheet("検索条件")
    meta.append(["項目", "値"])
    meta.append(["取得種別", "発注情報" if params.info_type == "order" else "契約結果情報"])
    meta.append(["取得元", "秋田県電子入札システム 発注情報" if params.info_type == "order" else "秋田県電子入札システム 契約結果情報"])
    meta.append(["契約年度", params.fiscal_year])
    meta.append(["調達区分", params.procurement_category_label])
    meta.append(["工事・委託名称キーワード", "\n".join(params.keywords) if params.keywords else "指定なし"])
    meta.append(["工事・委託場所", params.location or "指定なし"])
    meta.append(["請負者名", params.contractor or "指定なし"])
    meta.append(["取得件数", len(rows)])
    meta.append(["取得日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 80
    for cell in meta[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    wb.save(path)
    return path


class AppHandler(BaseHTTPRequestHandler):
    server_version = "AkitaProcurementApp/1.0"

    def log_message(self, fmt: str, *args) -> None:
        print("[%s] %s" % (self.log_date_time_string(), fmt % args))

    def _send_json(self, status: int, payload: dict) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _send_file(self, path: Path, download_name: str | None = None) -> None:
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        if download_name:
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(download_name)}")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        if self.path == "/":
            self._send_file(PUBLIC_DIR / "index.html")
            return
        if self.path.startswith("/output/"):
            name = posixpath.basename(unquote(self.path))
            self._send_file(OUTPUT_DIR / name)
            return

        rel = unquote(self.path.lstrip("/"))
        safe_path = (PUBLIC_DIR / rel).resolve()
        if PUBLIC_DIR.resolve() in safe_path.parents or safe_path == PUBLIC_DIR.resolve():
            self._send_file(safe_path)
            return
        self.send_error(404)

    def do_POST(self) -> None:
        if self.path != "/api/search":
            self.send_error(404)
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
            params = SearchParams.from_payload(payload)
            rows = scrape_contracts(params)
            workbook = write_workbook(rows, params)
            preview = rows[:30]
            self._send_json(
                200,
                {
                    "ok": True,
                    "count": len(rows),
                    "rows": preview,
                    "downloadUrl": f"/output/{workbook.name}",
                    "fileName": workbook.name,
                },
            )
        except Exception as exc:
            message = str(exc) or repr(exc)
            self._send_json(500, {"ok": False, "error": message})


def main() -> None:
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Akita procurement app: http://{host}:{port}")
    print("Ctrl+C to stop.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        threading.Thread(target=server.shutdown, daemon=True).start()
        time.sleep(0.1)


if __name__ == "__main__":
    main()
